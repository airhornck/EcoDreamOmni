"""Document parser — PDF / Word / Excel text extraction.

Aligned with PRD FUNC-2 "PDF/Word/Excel知识导入".
"""

import csv
import io
from pathlib import Path
from typing import Dict, List


def parse_pdf(file_path: str) -> List[Dict[str, any]]:
    """Extract text from each page of a PDF.

    Returns:
        [{"name": str, "content": str, "source_page": int}, ...]
    """
    from PyPDF2 import PdfReader

    reader = PdfReader(file_path)
    results = []
    base_name = Path(file_path).stem
    for i, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        text = text.strip()
        if text:
            results.append(
                {
                    "name": f"{base_name} 第{i}页",
                    "content": text,
                    "source_page": i,
                }
            )
    return results


def parse_docx(file_path: str) -> List[Dict[str, any]]:
    """Extract text from a Word document.

    Returns the whole document as a single entry.
    """
    from docx import Document

    doc = Document(file_path)
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    text = "\n".join(paragraphs)
    base_name = Path(file_path).stem
    if text:
        return [
            {
                "name": base_name,
                "content": text,
                "source_page": 1,
            }
        ]
    return []


def parse_xlsx(file_path: str) -> List[Dict[str, any]]:
    """Extract rows from an Excel workbook (first sheet only).

    Treats the first row as headers. Each subsequent row becomes an entry
    with name = first column value and content = JSON-like key:value text.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    base_name = Path(file_path).stem
    results = []
    for idx, row in enumerate(rows[1:], start=2):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        # name = first non-empty column, fallback to row index
        name = cells[0] if cells[0] else f"{base_name} 行{idx}"
        # content = all columns as "header: value" lines
        content_lines = []
        for h, c in zip(headers[1:], cells[1:]):
            if c:
                content_lines.append(f"{h}: {c}")
        content = "\n".join(content_lines)
        if not content:
            content = cells[0]
        results.append(
            {
                "name": name,
                "content": content,
                "source_page": idx,
            }
        )
    return results


def parse_csv_text(text: str) -> List[Dict[str, any]]:
    """Parse CSV text (already decoded).

    Expected columns: entry_type, name, content
    Optional columns: product_id, approval_number, sku_code, brand_name,
                      prohibited_claims, required_disclaimers
    """
    reader = csv.DictReader(io.StringIO(text))
    fieldnames = set(reader.fieldnames or [])
    required = {"entry_type", "name", "content"}
    if not required.issubset(fieldnames):
        raise ValueError(f"CSV must contain columns: {required}")

    results = []
    for row in reader:
        item = {
            "entry_type": row["entry_type"].strip(),
            "name": row["name"].strip(),
            "content": row["content"].strip(),
        }
        for optional in (
            "product_id",
            "approval_number",
            "sku_code",
            "brand_name",
        ):
            if optional in fieldnames:
                val = row.get(optional, "").strip()
                if val:
                    item[optional] = val
        # JSON list columns — semicolon or comma separated
        for list_col in ("prohibited_claims", "required_disclaimers"):
            if list_col in fieldnames:
                val = row.get(list_col, "").strip()
                if val:
                    delim = ";" if ";" in val else ","
                    item[list_col] = [v.strip() for v in val.split(delim) if v.strip()]
        results.append(item)
    return results


def parse_document(file_path: str) -> List[Dict[str, any]]:
    """Unified entry-point: dispatch by file extension."""
    ext = Path(file_path).suffix.lower()
    if ext == ".pdf":
        return parse_pdf(file_path)
    if ext in (".docx", ".doc"):
        return parse_docx(file_path)
    if ext in (".xlsx", ".xls"):
        return parse_xlsx(file_path)
    raise ValueError(f"Unsupported file extension: {ext}")


# ─── VetDrug official Excel parser ───

_VETDRUG_COLUMN_MAP = {
    # 中文列名 -> 英文字段名
    "批准文号": "approval_number",
    "通用名称": "generic_name",
    "通用名": "generic_name",
    "商品名称": "product_name",
    "商品名": "product_name",
    "生产企业": "manufacturer",
    "企业名称": "manufacturer",
    "生产地址": "manufacturer_address",
    "企业地址": "manufacturer_address",
    "地址": "manufacturer_address",
    "主要成分": "ingredients",
    "成分": "ingredients",
    "规格": "specifications",
    "适应症": "indications",
    "功能主治": "indications",
    "用法用量": "usage_dosage",
    "用法与用量": "usage_dosage",
    "禁忌": "contraindications",
    "不良反应": "adverse_reactions",
    "注意事项": "precautions",
    "药物相互作用": "drug_interactions",
    "相互作用": "drug_interactions",
    "贮藏条件": "storage_conditions",
    "贮藏": "storage_conditions",
    "类别": "category",
    "分类": "category",
    "药品分类": "category",
    "处方药/非处方药": "drug_type",
    "药品类型": "drug_type",
    "有效期": "expiry_date",
    "有效期至": "expiry_date",
    "适用动物": "applicable_species",
    "适用物种": "applicable_species",
    "目标疾病": "target_diseases",
    "防治疾病": "target_diseases",
    # 英文列名 -> 自身
    "approval_number": "approval_number",
    "generic_name": "generic_name",
    "product_name": "product_name",
    "manufacturer": "manufacturer",
    "manufacturer_address": "manufacturer_address",
    "ingredients": "ingredients",
    "specifications": "specifications",
    "indications": "indications",
    "usage_dosage": "usage_dosage",
    "contraindications": "contraindications",
    "adverse_reactions": "adverse_reactions",
    "precautions": "precautions",
    "drug_interactions": "drug_interactions",
    "storage_conditions": "storage_conditions",
    "category": "category",
    "drug_type": "drug_type",
    "expiry_date": "expiry_date",
    "applicable_species": "applicable_species",
    "target_diseases": "target_diseases",
}


def parse_vetdrug_excel(file_path: str) -> List[Dict[str, any]]:
    """Parse official veterinary drug Excel from China Vet Drug Info Network.

    Supports both Chinese and English column headers.
    Returns list of dicts matching vetdrug_db_function.create_drug kwargs.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    raw_headers = [str(h).strip() if h else "" for h in rows[0]]
    field_map = {}
    for idx, h in enumerate(raw_headers):
        if not h:
            continue
        if h in _VETDRUG_COLUMN_MAP:
            field_map[idx] = _VETDRUG_COLUMN_MAP[h]
        # also try lower-cased / spaced variants
        elif h.lower().replace(" ", "_") in _VETDRUG_COLUMN_MAP:
            field_map[idx] = _VETDRUG_COLUMN_MAP[h.lower().replace(" ", "_")]

    if "approval_number" not in field_map.values():
        raise ValueError(
            "Excel must contain 'approval_number' column (批准文号). "
            f"Detected headers: {raw_headers}"
        )

    results = []
    for row in rows[1:]:
        item: Dict[str, any] = {}
        for idx, cell in enumerate(row):
            if idx not in field_map:
                continue
            field = field_map[idx]
            val = str(cell).strip() if cell is not None else ""
            if not val or val == "None":
                continue
            if field in ("applicable_species", "target_diseases", "tags"):
                delim = ";" if ";" in val else ","
                item[field] = [v.strip() for v in val.split(delim) if v.strip()]
            else:
                item[field] = val
        if item.get("approval_number"):
            results.append(item)
    return results
